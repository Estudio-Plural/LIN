"""Arma el set de validación humana de la clasificación (Fase 1).

Genera dos archivos:
  - LIN_validacion_fase1.xlsx  -> para enviar a Camino (anotación CIEGA).
  - CLAVE_fase1.csv            -> queda local; mapea id -> etiqueta del modelo
                                 para medir acuerdo humano-modelo después.

Método (no negociar): muestra ESTRATIFICADA por etiqueta y CARGADA a los casos
de frontera (menor `confianza` del modelo, donde más duda). La anotación es ciega
(no mostramos la etiqueta del modelo) para no anclar al anotador y poder medir el
acuerdo de verdad. Determinista: ordena por (preferencia hispana, confianza, id).

Uso:  python -m src.analyze.build_validation_set
Para Fase 2: subir los cupos en STRATA.
"""
from __future__ import annotations

import ast
import csv
import json
import os
from collections import defaultdict

VIDEOS = "data/dashboard/videos.json"
CLASSIFIED = "data/processed/classified.json"  # trae `confianza` por id
OUT_DIR = "data/processed/validacion_fase1"

# (etiqueta interna, texto amable para el menú desplegable)
LABELS = [
    ("manosfera_sincera", "Manosfera sincera — promueve la idea"),
    ("contra_critica", "Contra-crítica — la cuestiona o critica"),
    ("satira_humor", "Sátira / humor — parodia o ironía"),
    ("medio_periodismo", "Medio / periodismo — cubre el fenómeno"),
    ("falso_positivo", "Falso positivo — no es del tema"),
    ("ambiguo", "Ambiguo — no se puede determinar"),
]
LABEL_FRIENDLY = dict(LABELS)

SUBTEMAS = [
    "hipergamia", "alto valor", "red pill", "alfa", "looksmaxxing",
    "antifeminismo", "fitness", "gaming", "otro / no aplica",
]

# Cupos por etiqueta. Cargado a la triada de frontera (sincera/ambiguo/crítica)
# y con sátira sobre-representada porque Camino reportó cuentas de parodia.
STRATA = {
    "manosfera_sincera": 8,
    "ambiguo": 7,
    "contra_critica": 5,
    "satira_humor": 5,
    "falso_positivo": 3,
    "medio_periodismo": 2,
}

DEFINITIONS = [
    ("Manosfera sincera",
     "Promueve genuinamente ideas como hipergamia, “hombre de alto valor”, "
     "red pill, alfa, antifeminismo o looksmaxxing."),
    ("Contra-crítica",
     "Cuestiona o critica esas ideas (mirada feminista, progresista o educativa)."),
    ("Sátira / humor",
     "Parodia, comedia o ironía explícita sobre esas ideas."),
    ("Medio / periodismo",
     "Cobertura mediática o periodística del fenómeno (reportaje, nota), no promoción."),
    ("Falso positivo",
     "La palabra clave apareció por casualidad; el video no trata del tema."),
    ("Ambiguo",
     "No se puede determinar con claridad a cuál de las anteriores pertenece."),
]


def _is_latam(v) -> bool:
    return str(v).strip().lower() in ("true", "1")


def _hispanic_pref(rec) -> int:
    return int(rec.get("lang") in ("es", "pt") or _is_latam(rec.get("is_latam")))


def _hashtags(raw) -> str:
    if isinstance(raw, list):
        tags = raw
    else:
        try:
            tags = ast.literal_eval(raw) if raw else []
        except (ValueError, SyntaxError):
            tags = []
    tags = [t for t in tags if t]
    return " ".join("#" + t.lstrip("#") for t in tags)


def load_rows():
    videos = json.load(open(VIDEOS, encoding="utf-8"))
    conf = json.load(open(CLASSIFIED, encoding="utf-8"))
    for v in videos:
        c = conf.get(str(v.get("id")), {})
        v["confianza"] = c.get("confianza")
    return videos


def select(rows):
    by_label = defaultdict(list)
    for r in rows:
        if r.get("label") in STRATA:
            by_label[r["label"]].append(r)

    picked = {}
    for label, n in STRATA.items():
        pool = by_label[label]
        # frontera primero: hispano preferido, luego menor confianza
        pool.sort(key=lambda r: (-_hispanic_pref(r),
                                 r["confianza"] if r["confianza"] is not None else 1.0,
                                 str(r["id"])))
        picked[label] = pool[:n]

    # intercalar por etiquetas (round-robin) para no agrupar tipos en la planilla
    order = [lbl for lbl, _ in LABELS if lbl in picked]
    out, i = [], 0
    while any(len(picked[l]) > i for l in order):
        for l in order:
            if len(picked[l]) > i:
                out.append(picked[l][i])
        i += 1
    return out


def build_xlsx(selected, path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()

    # ---- Hoja: Listas (oculta, fuente de los menús) ----
    lst = wb.active
    lst.title = "Listas"
    lst["A1"] = "categorias"
    lst["B1"] = "subtemas"
    lst["C1"] = "siono"
    for idx, (_, friendly) in enumerate(LABELS, start=2):
        lst[f"A{idx}"] = friendly
    for idx, s in enumerate(SUBTEMAS, start=2):
        lst[f"B{idx}"] = s
    for idx, s in enumerate(["Sí", "No", "No sé"], start=2):
        lst[f"C{idx}"] = s
    lst.sheet_state = "hidden"
    n_cat = len(LABELS) + 1
    n_sub = len(SUBTEMAS) + 1

    # ---- Hoja: Instrucciones ----
    ins = wb.create_sheet("Instrucciones")
    ins.column_dimensions["A"].width = 26
    ins.column_dimensions["B"].width = 95
    h = Font(bold=True, size=14)
    sub = Font(bold=True, size=11)
    wrap = Alignment(wrap_text=True, vertical="top")

    ins["A1"] = "Validación de la clasificación — Fase 1"
    ins["A1"].font = h
    lines = [
        ("", ""),
        ("Qué es", "Una muestra de ~30 videos para contrastar a mano la etiqueta "
                   "que hoy pone el modelo de IA. Toma ~20 minutos. Priorizamos a "
                   "propósito casos de frontera (los que caen entre “manosfera”, "
                   "“crítica” y “ambiguo”), que es donde el modelo más duda."),
        ("", ""),
        ("Cómo marcar", ""),
        ("  1", "Abre el enlace y MIRA el video. El modelo solo vio el texto; tu "
                "ventaja es ver el video y la cuenta."),
        ("  2", "En la hoja “Etiquetado”, elige UNA categoría en la columna "
                "“Tu etiqueta” (menú desplegable)."),
        ("  3", "Opcional: marca el subtema."),
        ("  4", "Si la CUENTA entera parece de parodia/sátira (aunque ese video "
                "suelto parezca serio), marca “Sí” en la columna correspondiente. "
                "Esto nos ayuda con algo que ustedes ya notaron."),
        ("  5", "Anota cualquier duda o caso raro en “Notas”."),
        ("", ""),
        ("Importante", "Cada persona llena su PROPIA copia, sin ver la del otro. "
                       "A propósito no mostramos la etiqueta del modelo: así medimos "
                       "el acuerdo de verdad (entre ustedes y con el modelo)."),
        ("", ""),
        ("Las 6 categorías", ""),
    ]
    row = 2
    for a, b in lines:
        ins[f"A{row}"] = a
        ins[f"B{row}"] = b
        ins[f"B{row}"].alignment = wrap
        if a in ("Cómo marcar", "Importante", "Las 6 categorías", "Qué es"):
            ins[f"A{row}"].font = sub
        row += 1
    for name, desc in DEFINITIONS:
        ins[f"A{row}"] = name
        ins[f"A{row}"].font = Font(bold=True)
        ins[f"A{row}"].alignment = wrap
        ins[f"B{row}"] = desc
        ins[f"B{row}"].alignment = wrap
        row += 1

    # ---- Hoja: Etiquetado ----
    ws = wb.create_sheet("Etiquetado")
    headers = [
        "#", "Enlace (abrir y ver)", "País", "Keyword",
        "Texto (lo que vio el modelo)", "Hashtags",
        "Tu etiqueta", "Subtema (opcional)",
        "¿La cuenta entera parece parodia? (opcional)", "Notas / dudas",
    ]
    widths = [4, 46, 6, 16, 60, 28, 34, 18, 26, 30]
    head_fill = PatternFill("solid", fgColor="1F2937")     # gris oscuro
    input_fill = PatternFill("solid", fgColor="DCFCE7")    # verde claro = aquí marcas
    input_cols = {7, 8, 9, 10}
    for c, (title, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(1, c, title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = input_fill if c in input_cols else head_fill
        if c in input_cols:
            cell.font = Font(bold=True, color="166534")
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 42

    link_font = Font(color="2563EB", underline="single")
    for i, r in enumerate(selected, start=1):
        row = i + 1
        ws.cell(row, 1, i)
        cell = ws.cell(row, 2, r.get("url"))
        if r.get("url"):
            cell.hyperlink = r["url"]
            cell.font = link_font
        ws.cell(row, 3, r.get("loc") or "")
        ws.cell(row, 4, r.get("keyword") or "")
        tcell = ws.cell(row, 5, (r.get("text") or "").strip())
        tcell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 6, _hashtags(r.get("hashtags"))).alignment = Alignment(
            wrap_text=True, vertical="top")
        for c in input_cols:
            ws.cell(row, c).fill = input_fill
        ws.row_dimensions[row].height = 56

    last = len(selected) + 1
    dv_cat = DataValidation(type="list", formula1=f"Listas!$A$2:$A${n_cat}", allow_blank=True)
    dv_sub = DataValidation(type="list", formula1=f"Listas!$B$2:$B${n_sub}", allow_blank=True)
    dv_yn = DataValidation(type="list", formula1="Listas!$C$2:$C$4", allow_blank=True)
    for dv in (dv_cat, dv_sub, dv_yn):
        ws.add_data_validation(dv)
    dv_cat.add(f"G2:G{last}")
    dv_sub.add(f"H2:H{last}")
    dv_yn.add(f"I2:I{last}")

    ws.freeze_panes = "G2"
    wb._sheets = [ins, ws, lst]  # Instrucciones primero
    wb.active = 0

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)


def build_clave(selected, path):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["#", "id", "url", "etiqueta_modelo",
                    "etiqueta_modelo_amable", "subtema_modelo", "confianza"])
        for i, r in enumerate(selected, start=1):
            w.writerow([i, r.get("id"), r.get("url"), r.get("label"),
                        LABEL_FRIENDLY.get(r.get("label"), r.get("label")),
                        r.get("subtema"), r.get("confianza")])


def main():
    rows = load_rows()
    selected = select(rows)
    xlsx = os.path.join(OUT_DIR, "LIN_validacion_fase1.xlsx")
    clave = os.path.join(OUT_DIR, "CLAVE_fase1.csv")
    build_xlsx(selected, xlsx)
    build_clave(selected, clave)

    from collections import Counter
    dist = Counter(r["label"] for r in selected)
    print(f"Seleccionados: {len(selected)} videos")
    for lbl, _ in LABELS:
        if lbl in dist:
            confs = [r["confianza"] for r in selected
                     if r["label"] == lbl and r["confianza"] is not None]
            avg = sum(confs) / len(confs) if confs else float("nan")
            print(f"  {lbl:18s} {dist[lbl]:2d}   confianza media {avg:.2f}")
    print(f"\nPlanilla (enviar): {xlsx}")
    print(f"Clave (local):     {clave}")


if __name__ == "__main__":
    main()
